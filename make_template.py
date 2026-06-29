"""
Generate bu_control_reference_template.xlsx
Includes all reference sheets + DEL_COSTS + TRANSPORT_MATRIX

Run:
    python3 make_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL   = PatternFill("solid", fgColor="1A1917")   # dark bg
HEADER_FONT   = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
TITLE_FILL    = PatternFill("solid", fgColor="2D4739")   # green-dark
TITLE_FONT    = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
EXAMPLE_FILL  = PatternFill("solid", fgColor="F5F5F0")   # light grey
NOTE_FONT     = Font(color="888888", italic=True, name="Calibri", size=9)
BORDER        = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)

def style_header(ws, row_num, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_title(ws, row_num, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = TITLE_FILL
        cell.font   = TITLE_FONT

def style_example(ws, row_num, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = EXAMPLE_FILL
        cell.font = Font(color="555555", italic=True, name="Calibri", size=10)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A3"):
    ws.freeze_panes = cell

# ── PRODUCTS ─────────────────────────────────────────────────
def sheet_products(wb):
    ws = wb.create_sheet("PRODUCTS")
    ws.cell(1,1,"PRODUCTS — Internal product list").fill = TITLE_FILL
    ws.cell(1,1).font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws.cell(2,1,"Code");  ws.cell(2,2,"Name");  ws.cell(2,3,"Unit")
    style_header(ws, 2, 3)
    ws.cell(3,1,"ESO");   ws.cell(3,2,"Olio di Soia Epossidato"); ws.cell(3,3,"MT")
    style_example(ws, 3, 3)
    set_col_widths(ws, [15,40,8])
    freeze(ws)

# ── SUPPLIERS ────────────────────────────────────────────────
def sheet_suppliers(wb):
    ws = wb.create_sheet("SUPPLIERS")
    ws.cell(1,1,"SUPPLIERS").fill = TITLE_FILL
    ws.cell(1,1).font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.cell(2,1,"Name"); ws.cell(2,2,"Country (2-letter ISO)")
    style_header(ws, 2, 2)
    ws.cell(3,1,"ACME Chemicals SA"); ws.cell(3,2,"AR")
    style_example(ws, 3, 2)
    set_col_widths(ws, [40,20])
    freeze(ws)

# ── CUSTOMERS ────────────────────────────────────────────────
def sheet_customers(wb):
    ws = wb.create_sheet("CUSTOMERS")
    ws.cell(1,1,"CUSTOMERS").fill = TITLE_FILL
    ws.cell(1,1).font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.cell(2,1,"Name"); ws.cell(2,2,"Country (2-letter ISO)")
    style_header(ws, 2, 2)
    ws.cell(3,1,"Cliente Srl"); ws.cell(3,2,"IT")
    style_example(ws, 3, 2)
    set_col_widths(ws, [40,20])
    freeze(ws)

# ── STORAGE ──────────────────────────────────────────────────
def sheet_storage(wb):
    ws = wb.create_sheet("STORAGE")
    ws.cell(1,1,"STORAGE FACILITIES").fill = TITLE_FILL
    ws.cell(1,1).font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws.cell(2,1,"Code"); ws.cell(2,2,"Name"); ws.cell(2,3,"Location/Address"); ws.cell(2,4,"Country")
    style_header(ws, 2, 4)
    ws.cell(3,1,"WH-IT-01"); ws.cell(3,2,"Silomar Mortara"); ws.cell(3,3,"Mortara (PV)"); ws.cell(3,4,"IT")
    style_example(ws, 3, 4)
    set_col_widths(ws, [15,30,30,10])
    freeze(ws)

# ── CARRIERS ─────────────────────────────────────────────────
def sheet_carriers(wb):
    ws = wb.create_sheet("CARRIERS")
    ws.cell(1,1,"CARRIERS").fill = TITLE_FILL
    ws.cell(1,1).font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.cell(2,1,"Code"); ws.cell(2,2,"Name")
    style_header(ws, 2, 2)
    ws.cell(3,1,"TRS01"); ws.cell(3,2,"Trasporti Rossi Srl")
    style_example(ws, 3, 2)
    set_col_widths(ws, [15,40])
    freeze(ws)

# ── LEASING ──────────────────────────────────────────────────
def sheet_leasing(wb):
    ws = wb.create_sheet("LEASING")
    ws.cell(1,1,"CONTAINER LEASING COMPANIES").fill = TITLE_FILL
    ws.cell(1,1).font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws.cell(2,1,"Name"); ws.cell(2,2,"Code prefix"); ws.cell(2,3,"Free days"); ws.cell(2,4,"Demurrage €/day")
    style_header(ws, 2, 4)
    ws.cell(3,1,"Den Hartogh"); ws.cell(3,2,"DHDU"); ws.cell(3,3,14); ws.cell(3,4,85.0)
    style_example(ws, 3, 4)
    set_col_widths(ws, [30,15,12,18])
    freeze(ws)

# ── LOCATIONS ────────────────────────────────────────────────
def sheet_locations(wb):
    ws = wb.create_sheet("LOCATIONS")
    ws.cell(1,1,"LOCATIONS (depots, terminals, customer sites)").fill = TITLE_FILL
    ws.cell(1,1).font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws.cell(2,1,"Code"); ws.cell(2,2,"Name"); ws.cell(2,3,"Type"); ws.cell(2,4,"City"); ws.cell(2,5,"Country")
    style_header(ws, 2, 5)
    ws.cell(3,1,"GE-SILOMAR"); ws.cell(3,2,"Silomar Mortara"); ws.cell(3,3,"DEPOT"); ws.cell(3,4,"Mortara"); ws.cell(3,5,"IT")
    note = ws.cell(3,3)
    note.comment = None
    style_example(ws, 3, 5)
    # Type note
    ws.cell(4,1,"→ Type values:"); ws.cell(4,2,"DEPOT  |  SUPPLIER  |  CUSTOMER  |  PORT")
    ws.cell(4,1).font = NOTE_FONT; ws.cell(4,2).font = NOTE_FONT
    set_col_widths(ws, [18,35,12,20,10])
    freeze(ws)

# ── DEL_COSTS ────────────────────────────────────────────────
def sheet_del_costs(wb):
    ws = wb.create_sheet("DEL_COSTS")
    ws.cell(1,1,"DELIVERY COSTS — Purchase-side cost stack per lane  (filled by LI)").fill = TITLE_FILL
    ws.cell(1,1).font = TITLE_FONT
    ws.merge_cells("A1:K1")

    headers = [
        "Product Code",       # A
        "Supplier Name",      # B
        "Country (paese)",    # C
        "Valid From",         # D  ISO date YYYY-MM-DD
        "T1/T2",              # E  customs regime (T1=non-EU, T2=EU)
        "Log In €/MT",        # F  inbound logistics proxy
        "Commission €/MT",    # G
        "Porto €/MT",         # H  port cost
        "Dazio %",            # I  customs duty — PCT mode (0 for T2)
        "Dazio €/MT",         # J  customs duty — EUR_MT mode (use one OR the other)
        "Log ITA €/ctr",      # K  Italian logistics per container
        "Stoccaggio €/MT",    # L  storage cost flat per MT
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
    style_header(ws, 2, len(headers))

    # Example row
    ex = ["ESO", "ACME Chemicals SA", "AR", "2026-01-01", "T1",
          15.0, 3.0, None, 6.5, None, 180.0, 8.0]
    for col, v in enumerate(ex, 1):
        ws.cell(3, col, v)
    style_example(ws, 3, len(headers))

    # Notes row
    notes = [
        "Must match Products sheet code",
        "Must match Suppliers sheet name (or leave blank = all suppliers)",
        "2-letter ISO (or blank = all countries)",
        "YYYY-MM-DD — new row for each rate change",
        "T1 = non-EU origin (duty may apply) | T2 = EU origin (duty = 0)",
        "Inbound freight proxy €/MT (FOB/EXW)",
        "Trading commission €/MT",
        "Port handling €/MT (if applicable)",
        "Duty as % of purchase price — e.g. 6.5 means 6.5% | 0 for T2",
        "Duty as flat €/MT — fill EITHER col I or col J, not both",
        "Italian road transport to depot €/container",
        "Storage cost €/MT (flat, one-time)",
    ]
    for col, n in enumerate(notes, 1):
        c = ws.cell(4, col, n)
        c.font = NOTE_FONT
        c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 36

    set_col_widths(ws, [16, 30, 14, 14, 8, 12, 14, 10, 10, 10, 14, 14])
    freeze(ws)

# ── TRANSPORT_MATRIX ─────────────────────────────────────────
def sheet_transport(wb):
    ws = wb.create_sheet("TRANSPORT_MATRIX")
    ws.cell(1,1,"TRANSPORT MATRIX — Outbound delivery rates  (filled by LI)").fill = TITLE_FILL
    ws.cell(1,1).font = TITLE_FONT
    ws.merge_cells("A1:G1")

    headers = [
        "Storage Facility Code",  # A
        "Destination",            # B  city or zone
        "Carrier Code",           # C
        "Product Code",           # D  blank = all products
        "Rate €/MT",              # E
        "Valid From",             # F
        "Valid To",               # G  blank = still active
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h)
    style_header(ws, 2, len(headers))

    ex = ["WH-IT-01", "MORTARA", "TRS01", "ESO", 12.5, "2026-01-01", None]
    for col, v in enumerate(ex, 1):
        ws.cell(3, col, v)
    style_example(ws, 3, len(headers))

    notes = [
        "Must match Storage sheet code",
        "City or zone — must match destination entered on sale",
        "Must match Carriers sheet code",
        "Product code or leave blank for all products",
        "€/MT outbound transport",
        "YYYY-MM-DD",
        "YYYY-MM-DD or leave blank if still active",
    ]
    for col, n in enumerate(notes, 1):
        c = ws.cell(4, col, n)
        c.font = NOTE_FONT
    set_col_widths(ws, [22, 20, 15, 16, 10, 14, 12])
    freeze(ws)


def build():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheet_products(wb)
    sheet_suppliers(wb)
    sheet_customers(wb)
    sheet_storage(wb)
    sheet_carriers(wb)
    sheet_leasing(wb)
    sheet_locations(wb)
    sheet_del_costs(wb)
    sheet_transport(wb)

    out = "bu_control_reference_template.xlsx"
    wb.save(out)
    print(f"✅ Template saved: {out}")
    print("   Sheets: " + ", ".join(wb.sheetnames))


if __name__ == "__main__":
    build()
