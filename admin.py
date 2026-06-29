"""
BU Control — Admin / Settings Blueprint
Manages all reference data: products, suppliers, customers,
storage facilities, carriers, supplier-depot defaults
"""

from flask import Blueprint, render_template, request, redirect, url_for, session, flash, g
import sqlite3, os

admin_bp = Blueprint("admin", __name__)

DB_PATH = os.environ.get("DB_PATH", os.path.join(os.path.dirname(__file__), "db", "bu_control.db"))

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") not in ("CEO", "LOGISTICS_ADMIN"):
            flash("Admin access required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

# ── Main settings page ───────────────────────────────────────
@admin_bp.route("/settings")
@admin_required
def settings():
    db = get_db()
    products   = db.execute("SELECT * FROM products   ORDER BY code").fetchall()
    suppliers  = db.execute("SELECT * FROM suppliers  ORDER BY name").fetchall()
    customers  = db.execute("SELECT * FROM customers  ORDER BY name").fetchall()
    facilities = db.execute("SELECT * FROM storage_facilities ORDER BY code").fetchall()
    carriers   = db.execute("SELECT * FROM carriers   ORDER BY name").fetchall()
    depot_defaults = db.execute("""
        SELECT sd.*, s.name AS supplier_name, sf.code AS facility_code,
               p.code AS product_code
        FROM supplier_depot_defaults sd
        JOIN suppliers s        ON s.id  = sd.supplier_id
        JOIN storage_facilities sf ON sf.id = sd.storage_facility_id
        LEFT JOIN products p    ON p.id  = sd.product_id
        ORDER BY s.name
    """).fetchall()
    aliases = db.execute("""
        SELECT pa.*, p.code AS product_code, s.name AS supplier_name
        FROM product_aliases pa
        JOIN products p   ON p.id  = pa.product_id
        JOIN suppliers s  ON s.id  = pa.supplier_id
        ORDER BY s.name, pa.alias
    """).fetchall()
    supplier_aliases = db.execute("""
        SELECT sa.*, s.name AS supplier_name
        FROM supplier_aliases sa
        JOIN suppliers s ON s.id = sa.supplier_id
        ORDER BY s.name, sa.alias
    """).fetchall()
    leasing_cos = db.execute("""
        SELECT lc.*, GROUP_CONCAT(lr.valid_from || ':' || lr.free_days || 'd@' || lr.demurrage_rate_eur_day || '€') AS rates
        FROM container_leasing_companies lc
        LEFT JOIN leasing_rates lr ON lr.leasing_company_id = lc.id
        GROUP BY lc.id ORDER BY lc.name
    """).fetchall()
    leasing_rates = db.execute("""
        SELECT lr.*, lc.name AS company_name
        FROM leasing_rates lr
        JOIN container_leasing_companies lc ON lc.id = lr.leasing_company_id
        ORDER BY lc.name, lr.valid_from DESC
    """).fetchall()
    locations = db.execute("SELECT * FROM locations ORDER BY type, name").fetchall()
    road_rates = db.execute("""
        SELECT rr.*, ca.name AS carrier_name,
               fl.name AS from_name, tl.name AS to_name
        FROM road_transport_rates rr
        JOIN carriers ca ON ca.id = rr.carrier_id
        LEFT JOIN locations fl ON fl.id = rr.from_location_id
        LEFT JOIN locations tl ON tl.id = rr.to_location_id
        ORDER BY ca.name, rr.valid_from DESC
    """).fetchall()
    del_costs = db.execute("""
        SELECT dc.*, p.code AS product_code, s.name AS supplier_name
        FROM delivery_costs dc
        LEFT JOIN products p  ON p.id  = dc.product_id
        LEFT JOIN suppliers s ON s.id  = dc.supplier_id
        ORDER BY p.code, s.name, dc.valid_from DESC
    """).fetchall()
    return render_template("settings.html",
        products=products, suppliers=suppliers, customers=customers,
        facilities=facilities, carriers=carriers, depot_defaults=depot_defaults,
        aliases=aliases, supplier_aliases=supplier_aliases,
        leasing_cos=leasing_cos, leasing_rates=leasing_rates,
        locations=locations, road_rates=road_rates,
        del_costs=del_costs,
    )

# ── PRODUCTS ─────────────────────────────────────────────────
@admin_bp.route("/settings/products/add", methods=["POST"])
@admin_required
def add_product():
    db = get_db()
    try:
        db.execute(
            "INSERT INTO products (code, name, unit) VALUES (?,?,?)",
            (request.form["code"].strip().upper(),
             request.form["name"].strip(),
             request.form.get("unit", "MT"))
        )
        db.commit()
        flash(f"Product {request.form['code']} added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#products")

@admin_bp.route("/settings/products/<int:pid>/toggle", methods=["POST"])
@admin_required
def toggle_product(pid):
    db = get_db()
    db.execute("UPDATE products SET is_active = 1 - is_active WHERE id=?", (pid,))
    db.commit()
    return redirect(url_for("admin.settings") + "#products")

# ── SUPPLIERS ────────────────────────────────────────────────
@admin_bp.route("/settings/suppliers/add", methods=["POST"])
@admin_required
def add_supplier():
    db = get_db()
    try:
        db.execute(
            "INSERT INTO suppliers (name, country) VALUES (?,?)",
            (request.form["name"].strip(),
             request.form.get("country", "").strip().upper() or None)
        )
        db.commit()
        flash(f"Supplier {request.form['name']} added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#suppliers")

@admin_bp.route("/settings/suppliers/<int:sid>/toggle", methods=["POST"])
@admin_required
def toggle_supplier(sid):
    db = get_db()
    db.execute("UPDATE suppliers SET is_active = 1 - is_active WHERE id=?", (sid,))
    db.commit()
    return redirect(url_for("admin.settings") + "#suppliers")

# ── CUSTOMERS ────────────────────────────────────────────────
@admin_bp.route("/settings/customers/add", methods=["POST"])
@admin_required
def add_customer():
    db = get_db()
    try:
        db.execute(
            "INSERT INTO customers (name, country) VALUES (?,?)",
            (request.form["name"].strip(),
             request.form.get("country", "").strip().upper() or None)
        )
        db.commit()
        flash(f"Customer {request.form['name']} added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#customers")

@admin_bp.route("/settings/customers/<int:cid>/toggle", methods=["POST"])
@admin_required
def toggle_customer(cid):
    db = get_db()
    db.execute("UPDATE customers SET is_active = 1 - is_active WHERE id=?", (cid,))
    db.commit()
    return redirect(url_for("admin.settings") + "#customers")

# ── STORAGE FACILITIES ───────────────────────────────────────
@admin_bp.route("/settings/facilities/add", methods=["POST"])
@admin_required
def add_facility():
    db = get_db()
    try:
        db.execute(
            "INSERT INTO storage_facilities (code, name, location, country) VALUES (?,?,?,?)",
            (request.form["code"].strip().upper(),
             request.form["name"].strip(),
             request.form.get("location", "").strip() or None,
             request.form.get("country", "IT").strip().upper())
        )
        db.commit()
        flash(f"Facility {request.form['code']} added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#facilities")

@admin_bp.route("/settings/facilities/<int:fid>/toggle", methods=["POST"])
@admin_required
def toggle_facility(fid):
    db = get_db()
    db.execute("UPDATE storage_facilities SET is_active = 1 - is_active WHERE id=?", (fid,))
    db.commit()
    return redirect(url_for("admin.settings") + "#facilities")

# ── CARRIERS ─────────────────────────────────────────────────
@admin_bp.route("/settings/carriers/add", methods=["POST"])
@admin_required
def add_carrier():
    db = get_db()
    try:
        db.execute(
            "INSERT INTO carriers (code, name) VALUES (?,?)",
            (request.form["code"].strip().upper(),
             request.form["name"].strip())
        )
        db.commit()
        flash(f"Carrier {request.form['name']} added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#carriers")

@admin_bp.route("/settings/carriers/<int:cid>/toggle", methods=["POST"])
@admin_required
def toggle_carrier(cid):
    db = get_db()
    db.execute("UPDATE carriers SET is_active = 1 - is_active WHERE id=?", (cid,))
    db.commit()
    return redirect(url_for("admin.settings") + "#carriers")

# ── SUPPLIER → DEPOT DEFAULTS ────────────────────────────────
@admin_bp.route("/settings/depot_defaults/add", methods=["POST"])
@admin_required
def add_depot_default():
    db = get_db()
    try:
        db.execute("""
            INSERT INTO supplier_depot_defaults
                (supplier_id, storage_facility_id, product_id, valid_from)
            VALUES (?,?,?,?)
        """, (
            request.form["supplier_id"],
            request.form["storage_facility_id"],
            request.form.get("product_id") or None,
            request.form.get("valid_from") or "2000-01-01",
        ))
        db.commit()
        flash("Depot default added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#depot-defaults")

@admin_bp.route("/settings/depot_defaults/<int:did>/delete", methods=["POST"])
@admin_required
def delete_depot_default(did):
    db = get_db()
    db.execute("DELETE FROM supplier_depot_defaults WHERE id=?", (did,))
    db.commit()
    flash("Depot default removed.", "success")
    return redirect(url_for("admin.settings") + "#depot-defaults")

# ── PRODUCT ALIASES ──────────────────────────────────────────
@admin_bp.route("/settings/aliases/add", methods=["POST"])
@admin_required
def add_alias():
    db = get_db()
    try:
        db.execute("""
            INSERT OR IGNORE INTO product_aliases (product_id, supplier_id, alias, source)
            VALUES (?,?,?,'MANUAL')
        """, (
            request.form["product_id"],
            request.form["supplier_id"],
            request.form["alias"].strip(),
        ))
        db.commit()
        flash("Alias added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#aliases")

@admin_bp.route("/settings/aliases/<int:aid>/delete", methods=["POST"])
@admin_required
def delete_alias(aid):
    db = get_db()
    db.execute("DELETE FROM product_aliases WHERE id=?", (aid,))
    db.commit()
    flash("Alias removed.", "success")
    return redirect(url_for("admin.settings") + "#aliases")

# ── EDIT routes ──────────────────────────────────────────────
@admin_bp.route("/settings/suppliers/<int:sid>/edit", methods=["POST"])
@admin_required
def edit_supplier(sid):
    db = get_db()
    try:
        db.execute("""
            UPDATE suppliers SET name=?, country=?, customs_regime=?, updated_at=datetime('now')
            WHERE id=?
        """, (
            request.form["name"].strip(),
            request.form.get("country", "").strip().upper() or None,
            request.form.get("customs_regime", "").strip() or None,
            sid
        ))
        db.commit()
        flash("Supplier updated.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#suppliers")

@admin_bp.route("/settings/customers/<int:cid>/edit", methods=["POST"])
@admin_required
def edit_customer(cid):
    db = get_db()
    try:
        db.execute("""
            UPDATE customers SET name=?, country=?, updated_at=datetime('now')
            WHERE id=?
        """, (
            request.form["name"].strip(),
            request.form.get("country", "").strip().upper() or None,
            cid
        ))
        db.commit()
        flash("Customer updated.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#customers")

@admin_bp.route("/settings/products/<int:pid>/edit", methods=["POST"])
@admin_required
def edit_product(pid):
    db = get_db()
    try:
        db.execute("""
            UPDATE products SET code=?, name=?, unit=?, updated_at=datetime('now')
            WHERE id=?
        """, (
            request.form["code"].strip().upper(),
            request.form["name"].strip(),
            request.form.get("unit", "MT"),
            pid
        ))
        db.commit()
        flash("Product updated.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#products")

@admin_bp.route("/settings/facilities/<int:fid>/edit", methods=["POST"])
@admin_required
def edit_facility(fid):
    db = get_db()
    try:
        db.execute("""
            UPDATE storage_facilities
            SET code=?, name=?, location=?, country=?, updated_at=datetime('now')
            WHERE id=?
        """, (
            request.form["code"].strip().upper(),
            request.form["name"].strip(),
            request.form.get("location", "").strip() or None,
            request.form.get("country", "IT").strip().upper(),
            fid
        ))
        db.commit()
        flash("Facility updated.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#facilities")

# ── LEASING COMPANIES ────────────────────────────────────────
@admin_bp.route("/settings/leasing/add", methods=["POST"])
@admin_required
def add_leasing_company():
    db = get_db()
    try:
        db.execute(
            "INSERT INTO container_leasing_companies (name, code) VALUES (?,?)",
            (request.form["name"].strip(), request.form.get("code","").strip().upper() or None)
        )
        db.commit()
        flash(f"Leasing company added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#leasing")

@admin_bp.route("/settings/leasing/<int:lid>/edit", methods=["POST"])
@admin_required
def edit_leasing_company(lid):
    db = get_db()
    db.execute("UPDATE container_leasing_companies SET name=?, code=?, updated_at=datetime('now') WHERE id=?",
        (request.form["name"].strip(), request.form.get("code","").strip().upper() or None, lid))
    db.commit()
    flash("Updated.", "success")
    return redirect(url_for("admin.settings") + "#leasing")

@admin_bp.route("/settings/leasing/rate/add", methods=["POST"])
@admin_required
def add_leasing_rate(lid=None):
    db = get_db()
    lid = request.form.get("leasing_company_id")
    try:
        db.execute("""
            INSERT INTO leasing_rates (leasing_company_id, valid_from, free_days, demurrage_rate_eur_day, notes)
            VALUES (?,?,?,?,?)
        """, (lid, request.form["valid_from"], int(request.form["free_days"]),
              float(request.form["demurrage_rate_eur_day"]), request.form.get("notes","")))
        db.commit()
        flash("Rate added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#leasing")

# ── ROAD TRANSPORT RATES ─────────────────────────────────────
@admin_bp.route("/settings/road_rates/add", methods=["POST"])
@admin_required
def add_road_rate():
    db = get_db()
    try:
        db.execute("""
            INSERT INTO road_transport_rates
                (carrier_id, from_location_id, to_location_id, valid_from, rate_eur_per_mt, rate_eur_per_trip, notes)
            VALUES (?,?,?,?,?,?,?)
        """, (
            request.form["carrier_id"],
            request.form.get("from_location_id") or None,
            request.form.get("to_location_id") or None,
            request.form["valid_from"],
            float(request.form.get("rate_eur_per_mt") or 0) or None,
            float(request.form.get("rate_eur_per_trip") or 0) or None,
            request.form.get("notes","")
        ))
        db.commit()
        flash("Road rate added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#road-rates")

# ── LOCATIONS ────────────────────────────────────────────────
@admin_bp.route("/settings/locations/add", methods=["POST"])
@admin_required
def add_location():
    db = get_db()
    try:
        db.execute("""
            INSERT INTO locations (code, name, type, city, country)
            VALUES (?,?,?,?,?)
        """, (
            request.form.get("code","").strip().upper() or None,
            request.form["name"].strip(),
            request.form.get("type","OTHER"),
            request.form.get("city","").strip() or None,
            request.form.get("country","IT").strip().upper()
        ))
        db.commit()
        flash("Location added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#locations")

# ── REFERENCE DATA IMPORT (Excel) ────────────────────────────
@admin_bp.route("/settings/import-reference", methods=["GET", "POST"])
@admin_required
def import_reference():
    from flask import send_file
    import openpyxl, io, os

    if request.method == "GET":
        template_path = os.path.join(os.path.dirname(__file__), "bu_control_reference_template.xlsx")
        if os.path.exists(template_path):
            return send_file(template_path,
                as_attachment=True,
                download_name="bu_control_reference_template.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        flash("Template file not found.", "error")
        return redirect(url_for("admin.settings"))

    f = request.files.get("reference_file")
    if not f or not f.filename.endswith((".xlsx", ".xls")):
        flash("Please upload an Excel file (.xlsx).", "error")
        return redirect(url_for("admin.settings") + "#import")

    db = get_db()
    wb = openpyxl.load_workbook(f, data_only=True)
    counts = {}

    # PRODUCTS
    if "PRODUCTS" in wb.sheetnames:
        ws = wb["PRODUCTS"]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            code = str(row[0].value).strip().upper() if row[0].value else None
            name = str(row[1].value).strip() if row[1].value else None
            unit = str(row[2].value).strip() if row[2].value else "MT"
            if code and name:
                db.execute("INSERT OR IGNORE INTO products (code, name, unit) VALUES (?,?,?)",
                    (code, name, unit))
                n += 1
        counts["products"] = n

    # SUPPLIERS
    if "SUPPLIERS" in wb.sheetnames:
        ws = wb["SUPPLIERS"]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            name = str(row[0].value).strip() if row[0].value else None
            country = str(row[1].value).strip().upper() if row[1].value else None
            if name:
                db.execute("INSERT OR IGNORE INTO suppliers (name, country) VALUES (?,?)",
                    (name, country))
                n += 1
        counts["suppliers"] = n

    # CUSTOMERS
    if "CUSTOMERS" in wb.sheetnames:
        ws = wb["CUSTOMERS"]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            name = str(row[0].value).strip() if row[0].value else None
            country = str(row[1].value).strip().upper() if row[1].value else None
            if name:
                db.execute("INSERT OR IGNORE INTO customers (name, country) VALUES (?,?)",
                    (name, country))
                n += 1
        counts["customers"] = n

    # STORAGE
    if "STORAGE" in wb.sheetnames:
        ws = wb["STORAGE"]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            code = str(row[0].value).strip().upper() if row[0].value else None
            name = str(row[1].value).strip() if row[1].value else None
            location = str(row[2].value).strip() if row[2].value else None
            country = str(row[3].value).strip().upper() if row[3].value else "IT"
            if code and name:
                db.execute("""INSERT OR IGNORE INTO storage_facilities
                    (code, name, location, country) VALUES (?,?,?,?)""",
                    (code, name, location, country))
                n += 1
        counts["facilities"] = n

    # CARRIERS
    if "CARRIERS" in wb.sheetnames:
        ws = wb["CARRIERS"]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            code = str(row[0].value).strip().upper() if row[0].value else None
            name = str(row[1].value).strip() if row[1].value else None
            if name:
                db.execute("INSERT OR IGNORE INTO carriers (code, name) VALUES (?,?)",
                    (code, name))
                n += 1
        counts["carriers"] = n

    # LEASING
    if "LEASING" in wb.sheetnames:
        ws = wb["LEASING"]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            name = str(row[0].value).strip() if row[0].value else None
            code = str(row[1].value).strip().upper() if row[1].value else None
            free_days = int(row[2].value) if row[2].value else 14
            rate = float(row[3].value) if row[3].value else None
            if name:
                db.execute("""INSERT OR IGNORE INTO container_leasing_companies
                    (name, code) VALUES (?,?)""", (name, code))
                db.commit()
                lc = db.execute("SELECT id FROM container_leasing_companies WHERE name=?",
                    (name,)).fetchone()
                if lc and rate:
                    db.execute("""INSERT OR IGNORE INTO leasing_rates
                        (leasing_company_id, valid_from, free_days, demurrage_rate_eur_day)
                        VALUES (?,?,?,?)""",
                        (lc["id"], "2020-01-01", free_days, rate))
                n += 1
        counts["leasing"] = n

    # LOCATIONS
    if "LOCATIONS" in wb.sheetnames:
        ws = wb["LOCATIONS"]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            code = str(row[0].value).strip().upper() if row[0].value else None
            name = str(row[1].value).strip() if row[1].value else None
            ltype = str(row[2].value).strip().upper() if row[2].value else "OTHER"
            city = str(row[3].value).strip() if row[3].value else None
            country = str(row[4].value).strip().upper() if row[4].value else "IT"
            if name:
                db.execute("""INSERT OR IGNORE INTO locations
                    (code, name, type, city, country) VALUES (?,?,?,?,?)""",
                    (code, name, ltype, city, country))
                n += 1
        counts["locations"] = n

    # DEL_COSTS
    if "DEL_COSTS" in wb.sheetnames:
        ws = wb["DEL_COSTS"]
        n = 0

        def safe_float(v):
            try:
                return float(v) if v is not None else None
            except (ValueError, TypeError):
                return None

        def safe_date(v):
            if v is None:
                return None
            s = str(v).strip()
            if " " in s:
                s = s.split(" ")[0]
            if len(s) == 10 and s[4] == "-":
                return s
            return None

        for row in ws.iter_rows(min_row=3, values_only=False):
            valid_from = safe_date(row[3].value)
            if not valid_from:
                continue  # skip example row, notes row, empty rows

            product_code       = str(row[0].value).strip().upper() if row[0].value else None
            supplier_name      = str(row[1].value).strip()         if row[1].value else None
            paese              = str(row[2].value).strip().upper() if row[2].value else None
            customs_regime_val = str(row[4].value).strip().upper() if row[4].value else None
            log_in             = safe_float(row[5].value)
            commission         = safe_float(row[6].value)
            porto              = safe_float(row[7].value)
            dazio_pct          = safe_float(row[8].value)
            dazio_eur          = safe_float(row[9].value)
            log_ita            = safe_float(row[10].value)
            stoccaggio         = safe_float(row[11].value)

            product_id = None
            if product_code:
                p = db.execute("SELECT id FROM products WHERE code=?", (product_code,)).fetchone()
                product_id = p["id"] if p else None

            supplier_id = None
            if supplier_name:
                s = db.execute("SELECT id FROM suppliers WHERE name=?", (supplier_name,)).fetchone()
                supplier_id = s["id"] if s else None
                if supplier_id and customs_regime_val in ("T1", "T2"):
                    db.execute("UPDATE suppliers SET customs_regime=? WHERE id=?",
                               (customs_regime_val, supplier_id))
                # Update supplier customs_regime if provided
                if supplier_id and customs_regime_val in ("T1", "T2"):
                    db.execute("UPDATE suppliers SET customs_regime=? WHERE id=?",
                               (customs_regime_val, supplier_id))

            if dazio_pct is not None:
                dazio = dazio_pct
                dazio_type = "PCT"
            elif dazio_eur is not None:
                dazio = dazio_eur
                dazio_type = "EUR_MT"
            else:
                dazio = None
                dazio_type = "PCT"

            # Check if row already exists — update if so, insert if not
            existing = db.execute("""
                SELECT id FROM delivery_costs
                WHERE (product_id IS ? OR (product_id IS NULL AND ? IS NULL))
                  AND (supplier_id IS ? OR (supplier_id IS NULL AND ? IS NULL))
                  AND (paese IS ? OR (paese IS NULL AND ? IS NULL))
                  AND valid_from = ?
            """, (product_id, product_id, supplier_id, supplier_id,
                  paese, paese, valid_from)).fetchone()

            if existing:
                db.execute("""
                    UPDATE delivery_costs SET
                        log_in=?, commission=?, porto=?,
                        dazio=?, dazio_type=?, log_ita=?, stoccaggio=?,
                        source='EXCEL_IMPORT', updated_at=datetime('now')
                    WHERE id=?
                """, (log_in, commission, porto,
                      dazio, dazio_type, log_ita, stoccaggio,
                      existing["id"]))
            else:
                db.execute("""
                    INSERT INTO delivery_costs
                        (product_id, supplier_id, paese, valid_from,
                         log_in, commission, porto,
                         dazio, dazio_type, log_ita, stoccaggio, source)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    product_id, supplier_id, paese, valid_from,
                    log_in, commission, porto,
                    dazio, dazio_type, log_ita, stoccaggio,
                    "EXCEL_IMPORT"
                ))
            n += 1
        counts["del_costs"] = n

    # TRANSPORT_MATRIX
    if "TRANSPORT_MATRIX" in wb.sheetnames:
        ws = wb["TRANSPORT_MATRIX"]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            facility_code = str(row[0].value).strip().upper() if row[0].value else None
            destination   = str(row[1].value).strip()         if row[1].value else None
            carrier_code  = str(row[2].value).strip().upper() if row[2].value else None
            product_code  = str(row[3].value).strip().upper() if row[3].value else None
            try:
                rate = float(row[4].value) if row[4].value is not None else None
            except (ValueError, TypeError):
                rate = None
            try:
                valid_from = str(row[5].value).strip() if row[5].value else None
                if valid_from and len(valid_from) == 10 and valid_from[4] == "-":
                    pass
                else:
                    valid_from = None
            except Exception:
                valid_from = None
            valid_to = str(row[6].value).strip() if row[6].value else None

            if not (facility_code and destination and rate and valid_from):
                continue

            facility = db.execute(
                "SELECT id FROM storage_facilities WHERE code=?", (facility_code,)
            ).fetchone()
            carrier = db.execute(
                "SELECT id FROM carriers WHERE code=?", (carrier_code,)
            ).fetchone() if carrier_code else None
            product = db.execute(
                "SELECT id FROM products WHERE code=?", (product_code,)
            ).fetchone() if product_code else None

            if not facility or not carrier:
                continue

            db.execute("""
                INSERT INTO transport_matrix
                    (storage_facility_id, destination, carrier_id,
                     product_id, rate_eur_per_mt, valid_from, valid_to)
                VALUES (?,?,?,?,?,?,?)
            """, (
                facility["id"], destination, carrier["id"],
                product["id"] if product else None,
                rate, valid_from, valid_to or None,
            ))
            n += 1
        counts["transport_matrix"] = n

    db.commit()
    summary = ", ".join(f"{v} {k}" for k, v in counts.items() if v > 0)
    flash(f"Reference data imported: {summary}.", "success")
    return redirect(url_for("admin.settings"))

# ── SUPPLIER ALIASES ─────────────────────────────────────────
@admin_bp.route("/settings/supplier_aliases/add", methods=["POST"])
@admin_required
def add_supplier_alias():
    db = get_db()
    try:
        db.execute("""
            INSERT OR IGNORE INTO supplier_aliases (supplier_id, alias, source)
            VALUES (?, ?, 'MANUAL')
        """, (request.form["supplier_id"], request.form["alias"].strip()))
        db.commit()
        flash("Supplier alias added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("admin.settings") + "#supplier-aliases")

@admin_bp.route("/settings/supplier_aliases/<int:aid>/delete", methods=["POST"])
@admin_required
def delete_supplier_alias(aid):
    db = get_db()
    db.execute("DELETE FROM supplier_aliases WHERE id=?", (aid,))
    db.commit()
    flash("Alias removed.", "success")
    return redirect(url_for("admin.settings") + "#supplier-aliases")
